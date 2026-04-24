from flask import Flask, jsonify, request, send_from_directory, session
import openpyxl
import json
import os
import secrets
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, date
from functools import wraps

from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__, static_folder='.')

# Trust Railway's reverse proxy (fixes https:// detection and correct IPs)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Use a stable secret key from env — MUST be set in Railway env vars
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# ── Session cookie config ────────────────────────────────────────────────────
# Railway serves over HTTPS — cookies must be Secure + SameSite=Lax
# otherwise the browser silently drops the session cookie after login.
IS_PRODUCTION = os.environ.get('RAILWAY_ENVIRONMENT') is not None
app.config.update(
    SESSION_COOKIE_SECURE   = IS_PRODUCTION,   # HTTPS only on Railway, HTTP ok locally
    SESSION_COOKIE_HTTPONLY = True,            # JS cannot read the cookie
    SESSION_COOKIE_SAMESITE = 'Lax',           # sent on normal navigation
    PERMANENT_SESSION_LIFETIME = 86400 * 7,    # 7 days
)

# ── Persistent data directory ────────────────────────────────────────────────
# On Railway, use /data (mounted volume) if it exists, otherwise fall back to
# the project folder (local dev).  All mutable files live here so they survive
# redeploys.
BASE_DIR  = os.path.dirname(__file__)
DATA_DIR  = os.environ.get('DATA_DIR', BASE_DIR)   # override via Railway env var
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_FILE   = os.path.join(DATA_DIR, 'CLP Plan Approval.xlsx')
WOOQER_FILE  = os.path.join(DATA_DIR, 'Wooqer Fill.xlsx')
ATTEND_FILE  = os.path.join(DATA_DIR, 'Attendence.xlsx')
DATA_FILE    = os.path.join(DATA_DIR, 'data.json')
CONFIG_FILE  = os.path.join(DATA_DIR, 'email_config.json')

# Max upload size — 20 MB should cover any Excel file
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

ALLOWED_EXCEL = {'.xlsx', '.xls'}

ADMIN_USERNAME = 'admin'


# ══════════════════════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════════════════════

def load_email_config():
    # Priority: saved JSON file → Railway environment variables → defaults
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            try:
                cfg = json.load(f)
                if cfg.get('gmail') and cfg.get('app_password'):
                    return cfg
            except Exception:
                pass
    # Fall back to env vars (useful on Railway where file may not persist)
    return {
        'gmail':        os.environ.get('GMAIL_ADDRESS', ''),
        'app_password': os.environ.get('GMAIL_APP_PASSWORD', ''),
        'base_url':     os.environ.get('BASE_URL', 'http://localhost:5000'),
    }


def save_email_config(cfg):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {'requests': [], 'approvals': []}


def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2, default=str)


def read_visitors_from_excel():
    """Return dict  name → {name, designation, email, approverEmail, plans:[]}"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Sheet1']
    visitors = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cols = list(row) + [None] * 7
        designation, name, visit_date, plan, visitor_email, approver_email, change_plan = cols[:7]
        if not name or not visit_date:
            continue
        if isinstance(visit_date, datetime):
            date_str = visit_date.strftime('%Y-%m-%d')
        elif isinstance(visit_date, date):
            date_str = visit_date.strftime('%Y-%m-%d')
        else:
            date_str = str(visit_date)
        plan        = str(plan).strip()        if plan        else ''
        change_plan = str(change_plan).strip() if change_plan else ''
        if name not in visitors:
            visitors[name] = {
                'name':          name,
                'designation':   designation or '',
                'email':         (visitor_email or '').strip().lower(),
                'approverEmail': (approver_email or '').strip().lower(),
                'plans':         []
            }
        visitors[name]['plans'].append({'date': date_str, 'plan': plan, 'updatedPlan': change_plan})
    for v in visitors.values():
        v['plans'].sort(key=lambda x: x['date'])
    return visitors


def write_change_plan_to_excel(visitor_name, visit_date_str, new_plan):
    """Write new_plan into the 'Change Plan' column (col 7) for the matching
    visitor + date row in the Excel file. Returns (True, '') or (False, error)."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Sheet1']
        updated = 0
        for row in ws.iter_rows(min_row=2):
            cell_name = row[1].value   # col B — NAME
            cell_date = row[2].value   # col C — Visit Date
            if not cell_name or not cell_date:
                continue
            if cell_name != visitor_name:
                continue
            # Normalise date
            if isinstance(cell_date, (datetime, date)):
                row_date = cell_date.strftime('%Y-%m-%d') if isinstance(cell_date, datetime) else cell_date.strftime('%Y-%m-%d')
            else:
                row_date = str(cell_date)
            if row_date == visit_date_str:
                row[6].value = new_plan   # col G — Change Plan
                updated += 1
        if updated:
            wb.save(EXCEL_FILE)
            return True, f'Updated {updated} row(s) in Excel'
        return False, f'No matching row found for {visitor_name} on {visit_date_str}'
    except Exception as e:
        return False, str(e)


def _normalise_date(v):
    """Convert datetime / date / string to YYYY-MM-DD string."""
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    return str(v) if v else None


def read_wooqer_lookup():
    """Return dict keyed by (name_lower, 'YYYY-MM-DD') → store_code."""
    lookup = {}
    if not os.path.exists(WOOQER_FILE):
        return lookup
    try:
        wb = openpyxl.load_workbook(WOOQER_FILE, data_only=True)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2, values_only=True):
            dt, name, store, desig = (list(row) + [None]*4)[:4]
            d = _normalise_date(dt)
            if d and name and store and str(name).strip().lower() != '(blank)':
                lookup[(name.strip().lower(), d)] = str(store).strip().upper()
    except Exception:
        pass
    return lookup


def read_attendance_lookup():
    """Return dict keyed by (name_lower, 'YYYY-MM-DD') → {store, status}."""
    lookup = {}
    if not os.path.exists(ATTEND_FILE):
        return lookup
    try:
        wb = openpyxl.load_workbook(ATTEND_FILE, data_only=True)
        ws = wb['Attnd']
        for row in ws.iter_rows(min_row=2, values_only=True):
            cols = list(row) + [None]*20
            emp_name   = cols[14]   # EMP_NAME
            punch_date = cols[7]    # PUNCH_DATE2
            store      = cols[9]    # STORE_NAME
            status     = cols[13]   # STATUS  (P / A)
            if emp_name and punch_date and store and str(emp_name).strip() != '-':
                d = _normalise_date(punch_date)
                if d:
                    key = (emp_name.strip().lower(), d)
                    # Keep first punch record per person per day
                    if key not in lookup:
                        lookup[key] = {
                            'store':  str(store).strip().upper(),
                            'status': str(status).strip() if status else '—'
                        }
    except Exception:
        pass
    return lookup


# Plans that are not store visits — excluded from adherence %
NON_STORE_PLANS = {
    'leave', 'travelling', 'traveling', 'w/o', 'wo', 'w/off', 'off',
    'ho', 'new site', 'not in citykart', 'na', 'n/a', 'holiday',
    'weekly off', 'training', 'meeting'
}

def _is_store_visit(plan_str):
    """Return True if this plan entry is an actual store visit (not Leave/Off/etc.)."""
    return plan_str.strip().lower() not in NON_STORE_PLANS


def enrich_plans_with_compliance(visitors):
    """Add wooqer / attendance fields + adherence scores to every plan entry.

    adherence values:
      100  → visited correct store
        0  → wrong store or missing
     None  → N/A (Leave / Travelling / Off etc.)
    """
    wooqer_lk = read_wooqer_lookup()
    attend_lk  = read_attendance_lookup()
    for v in visitors.values():
        name_lower = v['name'].strip().lower()
        for p in v['plans']:
            d   = p['date']
            key = (name_lower, d)

            original_store = p['plan'].strip().upper()
            updated_store  = p['updatedPlan'].strip().upper() if p.get('updatedPlan') else ''
            effective      = updated_store or original_store
            is_visit       = _is_store_visit(p['plan'])

            # ── Wooqer ──────────────────────────────────────────────────────
            w_store = wooqer_lk.get(key)
            if w_store:
                w_match = (w_store == effective)
                p['wooqer'] = {'store': w_store, 'status': 'match' if w_match else 'diff'}
            else:
                w_match = False
                p['wooqer'] = {'store': '', 'status': 'missing'}

            p['wooqer_adherence'] = (100 if w_match else 0) if is_visit else None

            # ── Attendance vs ORIGINAL plan ──────────────────────────────────
            a_rec = attend_lk.get(key)
            if a_rec:
                a_store  = a_rec['store']
                a_present = (a_rec['status'] == 'P')
                ao_match  = a_present and (a_store == original_store)
                p['attendance_original'] = {
                    'store':  a_store,
                    'status': 'match' if ao_match else ('absent' if not a_present else 'diff')
                }
            else:
                ao_match = False
                p['attendance_original'] = {'store': '', 'status': 'missing'}

            p['attend_original_adherence'] = (100 if ao_match else 0) if is_visit else None

            # ── Attendance vs UPDATED plan ───────────────────────────────────
            if a_rec:
                a_store   = a_rec['store']
                a_present = (a_rec['status'] == 'P')
                au_match  = a_present and (a_store == effective)
                p['attendance_updated'] = {
                    'store':  a_store,
                    'status': 'match' if au_match else ('absent' if not a_present else 'diff'),
                    'has_update': bool(updated_store)
                }
            else:
                au_match = False
                p['attendance_updated'] = {
                    'store': '', 'status': 'missing', 'has_update': bool(updated_store)
                }

            p['attend_updated_adherence'] = (100 if au_match else 0) if is_visit else None

    return visitors


def get_current_user():
    """Return session dict or None."""
    return session.get('user')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_user():
            return jsonify({'success': False, 'error': 'Not logged in', 'code': 'UNAUTHORIZED'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'success': False, 'error': 'Not logged in', 'code': 'UNAUTHORIZED'}), 401
        if not user.get('isAdmin'):
            return jsonify({'success': False, 'error': 'Admin access required', 'code': 'FORBIDDEN'}), 403
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
#  Auth routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/login', methods=['POST'])
def login():
    body      = request.get_json() or {}
    identifier = body.get('identifier', '').strip()

    if not identifier:
        return jsonify({'success': False, 'error': 'Please enter your email or admin username'}), 400

    # ── Admin login ──────────────────────────────────────────────────────────
    if identifier.lower() == ADMIN_USERNAME:
        session.permanent = True
        session['user'] = {
            'name':    'Administrator',
            'email':   'admin',
            'isAdmin': True,
            'visitor': None
        }
        return jsonify({'success': True, 'user': session['user']})

    # ── Visitor OR Approver login by email ──────────────────────────────────
    email = identifier.lower()
    try:
        visitors = read_visitors_from_excel()
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cannot read Excel: {e}'}), 500

    # Check if visitor
    matched_visitor = next(
        (v for v in visitors.values() if v['email'].lower() == email), None
    )
    if matched_visitor:
        session.permanent = True
        session['user'] = {
            'name':       matched_visitor['name'],
            'email':      matched_visitor['email'],
            'isAdmin':    False,
            'isApprover': False,
            'visitor':    matched_visitor['name']
        }
        return jsonify({'success': True, 'user': session['user']})

    # Check if approver — collect all unique approver emails from Excel
    approver_emails = set(
        v['approverEmail'].lower()
        for v in visitors.values()
        if v.get('approverEmail')
    )
    if email in approver_emails:
        # Derive a display name from email (e.g. ritesh.rathi@citykart.org → Ritesh Rathi)
        local = email.split('@')[0]
        display_name = ' '.join(p.capitalize() for p in local.replace('.', ' ').split())
        session.permanent = True
        session['user'] = {
            'name':       display_name,
            'email':      email,
            'isAdmin':    False,
            'isApprover': True,
            'visitor':    None
        }
        return jsonify({'success': True, 'user': session['user']})

    return jsonify({'success': False,
                    'error': 'Email not found. Please use your registered CityKart email.'}), 404


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/me', methods=['GET'])
def me():
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'code': 'UNAUTHORIZED'}), 401
    return jsonify({'success': True, 'user': user})


# ══════════════════════════════════════════════════════════════════════════════
#  Visitor data
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/visitors', methods=['GET'])
@login_required
def get_visitors():
    try:
        visitors = read_visitors_from_excel()
        enrich_plans_with_compliance(visitors)   # add wooqer + attendance
        user = get_current_user()
        if user['isAdmin']:
            return jsonify({'success': True, 'visitors': visitors})
        # Visitor only gets their own data
        name = user['visitor']
        if name not in visitors:
            return jsonify({'success': True, 'visitors': {}})
        return jsonify({'success': True, 'visitors': {name: visitors[name]}})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
#  Change requests
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/requests', methods=['GET'])
@login_required
def get_requests():
    data = load_data()
    user = get_current_user()
    if user['isAdmin']:
        return jsonify({'success': True, 'requests': data['requests']})
    if user.get('isApprover'):
        # Approver sees requests assigned to their email
        mine = [r for r in data['requests'] if r.get('approverEmail','').lower() == user['email'].lower()]
        return jsonify({'success': True, 'requests': mine})
    # Visitor sees only their own requests
    mine = [r for r in data['requests'] if r['visitor'] == user['visitor']]
    return jsonify({'success': True, 'requests': mine})


@app.route('/api/requests', methods=['POST'])
@login_required
def create_request():
    user = get_current_user()
    body = request.get_json() or {}

    required = ['visitor', 'date', 'newPlan', 'reason', 'visitorEmail', 'approverEmail']
    for field in required:
        if not body.get(field):
            return jsonify({'success': False, 'error': f'Missing field: {field}'}), 400

    # Visitors can only submit for themselves
    if not user['isAdmin'] and body['visitor'] != user['visitor']:
        return jsonify({'success': False, 'error': 'You can only submit requests for yourself'}), 403

    data = load_data()
    change_req = {
        'id':            'CHG-' + str(int(datetime.now().timestamp() * 1000)),
        'visitor':       body['visitor'],
        'date':          body['date'],
        'newPlan':       body['newPlan'],
        'reason':        body['reason'],
        'visitorEmail':  body['visitorEmail'],
        'approverEmail': body['approverEmail'],
        'status':        'Pending',
        'timestamp':     datetime.now().isoformat(),
        'emailSent':     False,
        'emailError':    ''
    }

    # ── Save request FIRST — always succeeds regardless of email ────────────
    data['requests'].append(change_req)
    save_data(data)

    # ── Try sending email — failure does NOT block the request ───────────────
    # ok, msg = send_approval_email(change_req)   # temporarily commented out — mail issue under investigation
    ok, msg = False, 'Email disabled temporarily'

    change_req['emailSent']  = ok
    change_req['emailError'] = '' if ok else msg

    # Update saved record with email status
    for r in data['requests']:
        if r['id'] == change_req['id']:
            r['emailSent']  = ok
            r['emailError'] = change_req['emailError']
            break
    save_data(data)

    return jsonify({'success': True, 'request': change_req, 'emailSent': ok, 'emailMsg': msg}), 201


@app.route('/api/requests/<request_id>/approve', methods=['POST'])
@login_required
def approve_request_api(request_id):
    user = get_current_user()
    body = request.get_json() or {}
    comment = body.get('comment', '')

    # Only admin or assigned approver can approve
    if not user['isAdmin'] and not user.get('isApprover'):
        return jsonify({'success': False, 'error': 'Not authorised to approve requests'}), 403

    data   = load_data()
    change = next((r for r in data['requests'] if r['id'] == request_id), None)
    if not change:
        return jsonify({'success': False, 'error': 'Request not found'}), 404
    if change['status'] != 'Pending':
        return jsonify({'success': False, 'error': f'Already {change["status"]}'}), 400

    # Approver can only act on requests assigned to them
    if user.get('isApprover') and change.get('approverEmail','').lower() != user['email'].lower():
        return jsonify({'success': False, 'error': 'This request is not assigned to you'}), 403

    decided_by = 'admin' if user['isAdmin'] else user['name']
    original_plan = _get_original_plan(change)
    change['status'] = 'Approved'
    approval = _build_record(change, 'Approved', original_plan, comment or f'Approved by {decided_by}')
    data['approvals'].append(approval)
    save_data(data)
    # Write approved store code back to Excel
    write_change_plan_to_excel(change['visitor'], change['date'], change['newPlan'])
    return jsonify({'success': True, 'approval': approval})


@app.route('/api/requests/<request_id>/reject', methods=['POST'])
@login_required
def reject_request_api(request_id):
    user    = get_current_user()
    body    = request.get_json() or {}
    comment = body.get('comment', '').strip()

    # Only admin or assigned approver can reject
    if not user['isAdmin'] and not user.get('isApprover'):
        return jsonify({'success': False, 'error': 'Not authorised to reject requests'}), 403
    if not comment:
        return jsonify({'success': False, 'error': 'Comment is required for rejection'}), 400

    data   = load_data()
    change = next((r for r in data['requests'] if r['id'] == request_id), None)
    if not change:
        return jsonify({'success': False, 'error': 'Request not found'}), 404
    if change['status'] != 'Pending':
        return jsonify({'success': False, 'error': f'Already {change["status"]}'}), 400

    # Approver can only act on requests assigned to them
    if user.get('isApprover') and change.get('approverEmail','').lower() != user['email'].lower():
        return jsonify({'success': False, 'error': 'This request is not assigned to you'}), 403

    original_plan    = _get_original_plan(change)
    change['status'] = 'Rejected'
    rejection        = _build_record(change, 'Rejected', original_plan, comment)
    data['approvals'].append(rejection)
    save_data(data)
    return jsonify({'success': True, 'rejection': rejection})


# ══════════════════════════════════════════════════════════════════════════════
#  Approvals history
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/approvals', methods=['GET'])
@login_required
def get_approvals():
    data = load_data()
    user = get_current_user()
    if user['isAdmin']:
        return jsonify({'success': True, 'approvals': data['approvals']})
    if user.get('isApprover'):
        mine = [a for a in data['approvals'] if a.get('approverEmail','').lower() == user['email'].lower()]
        return jsonify({'success': True, 'approvals': mine})
    mine = [a for a in data['approvals'] if a['visitor'] == user['visitor']]
    return jsonify({'success': True, 'approvals': mine})


@app.route('/api/summary', methods=['GET'])
@admin_required
def get_summary():
    """Return per-visitor aggregated summary grouped by designation,
    optionally filtered by ?year=YYYY&month=M."""
    try:
        year  = request.args.get('year',  type=int)
        month = request.args.get('month', type=int)

        visitors_raw = read_visitors_from_excel()
        wooqer_lk    = read_wooqer_lookup()
        attend_lk    = read_attendance_lookup()

        # Designation sort order
        DESIG_ORDER = {
            'REGIONAL_MANAGER': 1,
            'CLUSTER_MANAGER':  2,
            'CLUSTER_LP':       3,
            'CLUSTER_CVM':      4,
        }

        rows = []
        for name, v in visitors_raw.items():
            plans = v['plans']
            if year and month:
                plans = [p for p in plans
                         if p['date'].startswith(f'{year}-{month:02d}')]
            if not plans:
                continue

            total = len(plans)
            store_days = plan_changed = wooqer_filled = wooqer_correct = 0
            att_correct_orig = att_correct_upd = att_present = 0

            for p in plans:
                plan_str   = p['plan'].strip()
                change_str = p.get('updatedPlan', '').strip()
                effective  = change_str.upper() if change_str else plan_str.upper()
                key        = (name.strip().lower(), p['date'])

                if not _is_store_visit(plan_str):
                    continue

                store_days += 1
                if change_str:
                    plan_changed += 1

                w = wooqer_lk.get(key, '')
                if w:
                    wooqer_filled += 1
                    if w == effective:
                        wooqer_correct += 1

                a = attend_lk.get(key, {})
                if a.get('status') == 'P':
                    att_present += 1
                    if a['store'] == plan_str.upper():
                        att_correct_orig += 1
                    if a['store'] == effective:
                        att_correct_upd += 1

            def pct(num, den):
                return round(num / den * 100) if den else None

            rows.append({
                'designation':         v['designation'] or '—',
                'name':                name,
                'total_days':          total,
                'store_days':          store_days,
                'plan_changed':        plan_changed,
                'plan_change_pct':     pct(plan_changed, store_days),
                'wooqer_filled':       wooqer_filled,
                'wooqer_adh_pct':      pct(wooqer_correct, store_days),
                'att_present':         att_present,
                'att_adh_orig_pct':    pct(att_correct_orig, store_days),
                'att_adh_upd_pct':     pct(att_correct_upd, store_days),
            })

        # Sort by designation hierarchy then name
        rows.sort(key=lambda r: (DESIG_ORDER.get(r['designation'], 99), r['name'].lower()))
        return jsonify({'success': True, 'summary': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/upload-excel', methods=['POST'])
@admin_required
def upload_excel():
    """Replace one of the three Excel source files.
    Form field 'file_type' must be: clp | wooqer | attendance
    """
    file_type = request.form.get('file_type', '').strip()
    targets = {
        'clp':        (EXCEL_FILE,  'CLP Plan Approval'),
        'wooqer':     (WOOQER_FILE, 'Wooqer Fill'),
        'attendance': (ATTEND_FILE, 'Attendance'),
    }
    if file_type not in targets:
        return jsonify({'success': False, 'error': 'Invalid file_type'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Empty filename'}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXCEL:
        return jsonify({'success': False, 'error': 'Only .xlsx / .xls files allowed'}), 400

    dest_path, label = targets[file_type]
    # Back up the old file just in case
    if os.path.exists(dest_path):
        backup = dest_path + '.bak'
        os.replace(dest_path, backup)
    try:
        f.save(dest_path)
        # Quick sanity check — make sure openpyxl can open it
        import openpyxl as _oxl
        _oxl.load_workbook(dest_path, read_only=True).close()
        return jsonify({'success': True, 'message': f'{label} updated successfully'})
    except Exception as e:
        # Restore backup on failure
        backup = dest_path + '.bak'
        if os.path.exists(backup):
            os.replace(backup, dest_path)
        return jsonify({'success': False, 'error': f'Invalid Excel file: {e}'}), 400


@app.route('/api/excel-status', methods=['GET'])
@admin_required
def excel_status():
    """Return last-modified timestamps for each Excel file."""
    def info(path):
        if os.path.exists(path):
            mtime = os.path.getmtime(path)
            return {
                'exists': True,
                'updated': datetime.fromtimestamp(mtime).strftime('%d %b %Y, %I:%M %p'),
                'size_kb': round(os.path.getsize(path) / 1024, 1)
            }
        return {'exists': False, 'updated': '—', 'size_kb': 0}

    return jsonify({
        'success':    True,
        'clp':        info(EXCEL_FILE),
        'wooqer':     info(WOOQER_FILE),
        'attendance': info(ATTEND_FILE),
    })


@app.route('/api/clear', methods=['POST'])
@admin_required
def clear_data():
    save_data({'requests': [], 'approvals': []})
    return jsonify({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
#  Email action links  (no login needed — approver clicks from email)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/action/approve/<request_id>', methods=['GET'])
def action_approve(request_id):
    data   = load_data()
    change = next((r for r in data['requests'] if r['id'] == request_id), None)
    if not change:
        return "<h2>❌ Request not found.</h2>", 404

    vd = datetime.strptime(change['date'], '%Y-%m-%d').strftime('%d %b %Y')
    if change['status'] != 'Pending':
        return confirmation_page(f'Already {change["status"]}', '⚠️', '#f59e0b',
                                 change['visitor'], vd, change['newPlan'])

    change['status'] = 'Approved'
    data['approvals'].append(
        _build_record(change, 'Approved', _get_original_plan(change), 'Approved via email link'))
    save_data(data)
    # Write approved store code back to Excel
    write_change_plan_to_excel(change['visitor'], change['date'], change['newPlan'])
    return confirmation_page('Request Approved!', '✅', '#10b981',
                             change['visitor'], vd, change['newPlan'])


@app.route('/action/reject/<request_id>', methods=['GET'])
def action_reject(request_id):
    data   = load_data()
    change = next((r for r in data['requests'] if r['id'] == request_id), None)
    if not change:
        return "<h2>❌ Request not found.</h2>", 404

    vd = datetime.strptime(change['date'], '%Y-%m-%d').strftime('%d %b %Y')
    if change['status'] != 'Pending':
        return confirmation_page(f'Already {change["status"]}', '⚠️', '#f59e0b',
                                 change['visitor'], vd, change['newPlan'])

    change['status'] = 'Rejected'
    data['approvals'].append(
        _build_record(change, 'Rejected', _get_original_plan(change), 'Rejected via email link'))
    save_data(data)
    return confirmation_page('Request Rejected', '❌', '#ef4444',
                             change['visitor'], vd, change['newPlan'])


# ══════════════════════════════════════════════════════════════════════════════
#  Email config
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/email-config', methods=['GET'])
@admin_required
def get_email_config():
    cfg = load_email_config()
    return jsonify({
        'success':    True,
        'gmail':      cfg.get('gmail', ''),
        'configured': bool(cfg.get('gmail') and cfg.get('app_password')),
        'base_url':   cfg.get('base_url', 'http://localhost:5000')
    })


@app.route('/api/email-config', methods=['POST'])
@admin_required
def save_email_config_api():
    body         = request.get_json() or {}
    gmail        = body.get('gmail', '').strip()
    app_password = body.get('app_password', '').strip()
    base_url     = body.get('base_url', '').strip()

    if not gmail or not app_password:
        return jsonify({'success': False, 'error': 'Gmail and App Password are required'}), 400

    # Auto-detect server IP if base_url is empty or still localhost
    if not base_url or base_url in ('http://localhost:5000', 'http://127.0.0.1:5000'):
        import socket
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
            s.close()
        except Exception:
            local_ip = 'localhost'
        base_url = f'http://{local_ip}:5000'

    save_email_config({'gmail': gmail, 'app_password': app_password, 'base_url': base_url})
    return jsonify({'success': True, 'message': 'Email configuration saved', 'base_url': base_url})


@app.route('/api/server-info', methods=['GET'])
def server_info():
    """Return the server's local network IP so admin can set base_url correctly."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = '127.0.0.1'
    return jsonify({'ip': local_ip, 'url': f'http://{local_ip}:5000'})


@app.route('/api/email-config/test', methods=['POST'])
@admin_required
def test_email():
    cfg = load_email_config()
    gmail        = cfg.get('gmail', '').strip()
    app_password = cfg.get('app_password', '').strip()
    if not gmail or not app_password:
        return jsonify({'success': False,
                        'error': f'Email not configured. gmail={bool(gmail)}, password={bool(app_password)}'}), 400
    fake_req = {
        'id': 'TEST-001', 'visitor': 'Test Visitor',
        'date': datetime.now().strftime('%Y-%m-%d'), 'newPlan': 'TEST',
        'reason': 'SMTP configuration test.',
        'visitorEmail': gmail, 'approverEmail': gmail
    }
    ok, msg = send_approval_email(fake_req)
    # Return full detail so admin can see exactly what failed
    return jsonify({'success': ok, 'message': msg,
                    'gmail': gmail, 'base_url': cfg.get('base_url','')})


# ══════════════════════════════════════════════════════════════════════════════
#  Static pages
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory('.', 'Visit_Plan_Approval.html')


# ══════════════════════════════════════════════════════════════════════════════
#  SMTP sender
# ══════════════════════════════════════════════════════════════════════════════

def send_approval_email(change_req):
    cfg          = load_email_config()
    gmail        = cfg.get('gmail', '').strip()
    app_password = cfg.get('app_password', '').strip()
    base_url     = cfg.get('base_url', 'http://localhost:5000').rstrip('/')

    if not gmail or not app_password:
        return False, 'Email not configured. Please save Gmail settings first.'

    req_id       = change_req['id']
    visitor_name = change_req['visitor']
    visit_date   = datetime.strptime(change_req['date'], '%Y-%m-%d').strftime('%d %b %Y')
    new_plan     = change_req['newPlan']
    reason       = change_req['reason']
    to_email     = change_req['approverEmail']
    from_email   = change_req['visitorEmail']
    approve_url  = f"{base_url}/action/approve/{req_id}"
    reject_url   = f"{base_url}/action/reject/{req_id}"

    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><style>
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f4f4f4;margin:0;padding:20px}}
  .c{{max-width:560px;margin:0 auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.08)}}
  .h{{background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;padding:28px 30px;text-align:center}}
  .h h2{{margin:0;font-size:20px}}.h p{{margin:6px 0 0;font-size:13px;opacity:.85}}
  .b{{padding:28px 30px}}
  .ib{{background:#f0f4ff;border-left:4px solid #667eea;border-radius:6px;padding:16px 18px;margin-bottom:22px}}
  .ir{{display:flex;margin-bottom:10px;font-size:14px}}.ir:last-child{{margin-bottom:0}}
  .il{{color:#666;min-width:110px;font-weight:600}}.iv{{color:#333}}
  .ac{{text-align:center;margin:26px 0 10px}}
  .btn{{display:inline-block;padding:14px 32px;border-radius:8px;font-size:15px;font-weight:700;text-decoration:none;margin:0 8px}}
  .ba{{background:#10b981;color:#fff}}.br{{background:#ef4444;color:#fff}}
  .f{{background:#f9f9f9;text-align:center;padding:14px;font-size:11px;color:#aaa;border-top:1px solid #eee}}
</style></head><body>
<div class="c">
  <div class="h"><h2>🗓️ Visit Plan Change Request</h2><p>Action required — please approve or reject</p></div>
  <div class="b">
    <p style="color:#555;font-size:14px;margin-bottom:20px"><strong>{visitor_name}</strong> has submitted a visit plan change request.</p>
    <div class="ib">
      <div class="ir"><span class="il">Visitor</span><span class="iv">{visitor_name}</span></div>
      <div class="ir"><span class="il">Visit Date</span><span class="iv">{visit_date}</span></div>
      <div class="ir"><span class="il">New Plan</span><span class="iv"><strong>{new_plan}</strong></span></div>
      <div class="ir"><span class="il">Reason</span><span class="iv">{reason}</span></div>
      <div class="ir"><span class="il">From</span><span class="iv">{from_email}</span></div>
    </div>
    <div class="ac">
      <a href="{approve_url}" class="btn ba">✅ Approve</a>
      <a href="{reject_url}"  class="btn br">❌ Reject</a>
    </div>
    <p style="text-align:center;font-size:12px;color:#aaa;margin-top:14px">No login required · ID: {req_id}</p>
  </div>
  <div class="f">CityKart Stores — Visit Plan Approval System</div>
</div></body></html>"""

    msg = MIMEMultipart('alternative')
    msg['Subject']  = f"[Action Required] Visit Plan Change — {visitor_name} on {visit_date}"
    msg['From']     = f"Visit Plan System <{gmail}>"
    msg['To']       = to_email
    msg['Reply-To'] = from_email
    msg.attach(MIMEText(html_body, 'html'))

    # Try port 465 (SSL) first — Railway blocks 587 (STARTTLS)
    # Fall back to 587 if 465 fails (for local dev compatibility)
    last_error = ''
    for method in ['ssl', 'starttls']:
        try:
            if method == 'ssl':
                import ssl as _ssl
                context = _ssl.create_default_context()
                with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as s:
                    s.login(gmail, app_password)
                    s.sendmail(gmail, to_email, msg.as_string())
            else:
                with smtplib.SMTP('smtp.gmail.com', 587) as s:
                    s.ehlo(); s.starttls(); s.login(gmail, app_password)
                    s.sendmail(gmail, to_email, msg.as_string())
            return True, f'Email sent successfully (via {method.upper()})'
        except smtplib.SMTPAuthenticationError:
            return False, 'Gmail authentication failed. Check your App Password.'
        except Exception as e:
            last_error = f'{method.upper()}: {e}'
            continue
    return False, f'Email failed. {last_error}'


# ══════════════════════════════════════════════════════════════════════════════
#  Confirmation page (email link result)
# ══════════════════════════════════════════════════════════════════════════════

def confirmation_page(title, icon, color, visitor, visit_date, new_plan):
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f4f4f4;display:flex;
       align-items:center;justify-content:center;min-height:100vh;padding:20px}}
  .card{{background:#fff;border-radius:14px;padding:44px 40px;max-width:460px;width:100%;
         text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.10)}}
  .icon{{font-size:64px;margin-bottom:16px}}
  h1{{font-size:26px;color:{color};margin-bottom:10px}}
  .sub{{color:#777;font-size:14px;margin-bottom:28px}}
  .info{{background:#f8f8f8;border-radius:8px;padding:16px 20px;text-align:left;margin-bottom:24px}}
  .row{{display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #eee;font-size:13px}}
  .row:last-child{{border-bottom:none}}
  .lbl{{color:#999}}.val{{color:#333;font-weight:600}}
  .back{{display:inline-block;margin-top:10px;padding:10px 24px;
         background:linear-gradient(135deg,#667eea,#764ba2);
         color:#fff;border-radius:7px;text-decoration:none;font-size:14px;font-weight:600}}
</style></head><body>
<div class="card">
  <div class="icon">{icon}</div>
  <h1>{title}</h1>
  <p class="sub">Your decision has been recorded.</p>
  <div class="info">
    <div class="row"><span class="lbl">Visitor</span><span class="val">{visitor}</span></div>
    <div class="row"><span class="lbl">Date</span><span class="val">{visit_date}</span></div>
    <div class="row"><span class="lbl">New Plan</span><span class="val">{new_plan}</span></div>
  </div>
  <a href="/" class="back">← Go to Dashboard</a>
</div></body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  Internal helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get_original_plan(change):
    try:
        visitors = read_visitors_from_excel()
        visitor  = visitors.get(change['visitor'], {})
        entry    = next((p for p in visitor.get('plans', []) if p['date'] == change['date']), None)
        return entry['plan'] if entry else ''
    except Exception:
        return ''


def _build_record(change, status, original_plan, comment):
    prefix = 'APR' if status == 'Approved' else 'REJ'
    ts_key = 'approvedAt' if status == 'Approved' else 'rejectedAt'
    return {
        'id':            f'{prefix}-{int(datetime.now().timestamp()*1000)}',
        'changeId':      change['id'],
        'visitor':       change['visitor'],
        'date':          change['date'],
        'originalPlan':  original_plan,
        'newPlan':       change['newPlan'],
        'status':        status,
        'comment':       comment,
        'visitorEmail':  change['visitorEmail'],
        'approverEmail': change['approverEmail'],
        ts_key:          datetime.now().isoformat()
    }


# ══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("=" * 55)
    print("  Visit Plan Approval System — Flask + Login + SMTP")
    print("=" * 55)
    cfg = load_email_config()
    print(f"  Email  : {'✅ Configured (' + cfg['gmail'] + ')' if cfg.get('gmail') else '⚠️  Not configured'}")
    print(f"  Excel  : {EXCEL_FILE}")
    print(f"  Data   : {DATA_FILE}")
    print(f"  URL    : http://localhost:5000")
    print(f"  Admin  : type 'admin' on the login page")
    print("=" * 55)
    app.run(host='0.0.0.0', port=5000, debug=True)
